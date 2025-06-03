import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import os
import time
from datetime import datetime
import openpyxl

# VERSION INFO
VERSION = "v1.8.1"
VERSION_DATE = "2025-06-02"
DEBUG_MODE = False

def safe_metric(metrics, key, default=0):
    """Safely get a metric value with a default if missing"""
    return metrics.get(key, default)

def format_metric(value, format_type='number'):
    """Format metric values consistently"""
    try:
        if format_type == 'number':
            return f"{value:,}"
        elif format_type == 'hours':
            return f"{value:,.1f}"
        elif format_type == 'percentage':
            return f"{value:.1f}%"
        return str(value)
    except (TypeError, ValueError):
        return "0"  # Safe default for invalid values

def get_sorting_strategies():
    """Define all sorting strategies for min/max optimization"""
    return [
        {"name": "Start Date (Early First)", "columns": ["Start Date", "SO Number"], "ascending": [True, True]},
        {"name": "Start Date (Late First)", "columns": ["Start Date", "SO Number"], "ascending": [False, True]},
        {"name": "Demand (Small First)", "columns": ["Demand", "Start Date"], "ascending": [True, True]},
        {"name": "Demand (Large First)", "columns": ["Demand", "Start Date"], "ascending": [False, True]},
        {"name": "Hours (Quick First)", "columns": ["Hours_Calc", "Start Date"], "ascending": [True, True]},
        {"name": "Hours (Long First)", "columns": ["Hours_Calc", "Start Date"], "ascending": [False, True]},
        {"name": "Part Number (A-Z)", "columns": ["Part", "Start Date"], "ascending": [True, True]},
        {"name": "Part Number (Z-A)", "columns": ["Part", "Start Date"], "ascending": [False, True]},
        {"name": "Planner (A-Z)", "columns": ["Planner", "Start Date"], "ascending": [True, True]},
        {"name": "Planner (Z-A)", "columns": ["Planner", "Start Date"], "ascending": [False, True]}
    ]

def build_stock_dictionary(df_ipis):
    """Build stock dict using IPIS as primary source (StockTally is just a summary of IPIS)"""
    stock = {}
    
    # Use IPIS as the authoritative source
    if not df_ipis.empty:
        df_ipis["PART_NO"] = df_ipis["PART_NO"].astype(str)
        ipis_stock = df_ipis.groupby("PART_NO")["Available Qty"].sum().to_dict()
        stock.update(ipis_stock)
    else:
        print("WARNING: IPIS sheet is empty - no stock data available!")
    
    # Note: StockTally is derived from IPIS, so no fallback needed
    return stock

def process_single_scenario(filepath, scenario_name, status_callback=None, scenario_num=1, total_scenarios=1, sorting_strategy=None):
    """Process a single scenario file and return results with live progress updates"""
    
    if status_callback:
        strategy_name = sorting_strategy["name"] if sorting_strategy else "Default"
        status_callback(f"📂 [Scenario {scenario_num}/{total_scenarios}] Loading sheets for {os.path.basename(filepath)} ({strategy_name})...")
    
    # Load the sheets we need
    df_main = pd.read_excel(filepath, sheet_name="Demand")
    df_struct = pd.read_excel(filepath, sheet_name="Planned Demand")
    df_component_demand = pd.read_excel(filepath, sheet_name="Component Demand")
    df_ipis = pd.read_excel(filepath, sheet_name="IPIS")
    df_hours = pd.read_excel(filepath, sheet_name="Hours")
    df_pos = pd.read_excel(filepath, sheet_name="POs")
    
    # Validate required columns in Demand sheet
    required_columns = ['SO No', 'Part No', 'Planner', 'Start Date', 'Rev Qty Due']
    missing_columns = [col for col in required_columns if col not in df_main.columns]
    if missing_columns:
        raise ValueError(f"Missing required columns in Demand sheet: {', '.join(missing_columns)}")
    
    # Rename columns to match the rest of the code's expectations
    df_main = df_main.rename(columns={
        'SO No': 'SO Number',
        'Part No': 'Part',
        'Rev Qty Due': 'Demand'
    })
    
    if status_callback:
        strategy_name = sorting_strategy["name"] if sorting_strategy else "Default"
        status_callback(f"🔁 [Scenario {scenario_num}/{total_scenarios}] Processing commitments ({strategy_name})...")
    
    # Build stock dictionary
    stock = build_stock_dictionary(df_ipis)

    # Build committed_components
    committed_components = {}
    committed_parts_count = 0
    total_committed_qty = 0

    if not df_component_demand.empty:
        df_component_demand["Component Part Number"] = df_component_demand["Component Part Number"].astype(str)
        committed_summary = df_component_demand.groupby("Component Part Number")["Component Qty Required"].sum()
        committed_components = committed_summary.to_dict()
        committed_parts_count = len(committed_components)
        total_committed_qty = sum(committed_components.values())

    # Adjust stock for existing commitments
    for component, committed_qty in committed_components.items():
        if component in stock:
            stock[component] = max(0, stock[component] - committed_qty)
        else:
            stock[component] = 0

    # ADD THIS DEBUG SECTION:
    if DEBUG_MODE:
        print(f"\n=== DEBUG: STOCK ANALYSIS ===")
        print(f"DEBUG: Stock for 585711 = {stock.get('585711', 'NOT FOUND')}")
        print(f"DEBUG: Committed for 585711 = {committed_components.get('585711', 'NOT FOUND')}")
        if '585711' in stock and '585711' in committed_components:
            initial_stock = stock.get('585711', 0) + committed_components.get('585711', 0)
            print(f"DEBUG: Initial stock before commitments = {initial_stock}")
            print(f"DEBUG: After subtracting commitments = {stock.get('585711', 0)}")
        
    # Build labor standards dictionary (unchanged)
    df_hours["PART_NO"] = df_hours["PART_NO"].astype(str)
    labor_standards = df_hours.groupby("PART_NO")["Hours per Unit"].sum().to_dict()
    
    if status_callback:
        strategy_name = sorting_strategy["name"] if sorting_strategy else "Default"
        status_callback(f"🔁 [Scenario {scenario_num}/{total_scenarios}] Building planned demand structures ({strategy_name})...")
    
    # Build planned demand structure - UPDATED FROM BOM LOGIC
    planned_demand = df_struct[df_struct["Component Part Number"].notna()].copy()
    planned_demand["SO Number"] = planned_demand["SO Number"].astype(str)
    planned_demand["Component Part Number"] = planned_demand["Component Part Number"].astype(str)
    
    # Pre-process main data (unchanged)
    df_main['Start Date'] = pd.to_datetime(df_main['Start Date'], errors='coerce')
    df_main["Part"] = df_main["Part"].astype(str)
    df_main["Planner"] = df_main["Planner"].fillna("UNKNOWN").astype(str)
    df_main["Demand"] = pd.to_numeric(df_main["Demand"], errors='coerce').fillna(0)
    
    # Calculate hours for sorting (unchanged)
    df_main["Hours_Calc"] = df_main.apply(lambda row: 
        labor_standards.get(str(row["Part"]), 0) * row["Demand"], axis=1)
    
    # Apply sorting strategy (unchanged)
    if sorting_strategy:
        # Handle missing values appropriately for each column type
        for col in sorting_strategy["columns"]:
            if col == "Start Date":
                # Put NaT (missing dates) at the end
                df_main = df_main.sort_values(sorting_strategy["columns"], 
                                            ascending=sorting_strategy["ascending"], 
                                            na_position='last')
            else:
                df_main = df_main.sort_values(sorting_strategy["columns"], 
                                            ascending=sorting_strategy["ascending"])
    else:
        # Default sorting (original behavior)
        df_main = df_main.sort_values(['Start Date', 'SO Number'], na_position='last')
    
    df_main = df_main.reset_index(drop=True)
    
    results = []
    processed = 0
    total = len(df_main)
    
    # Baseline estimate: ~0.15 seconds per order (conservative estimate)
    baseline_time_per_order = 0.15
    processing_start_time = time.time()

    # Process each order sequentially with FREQUENT UI updates + TIME ESTIMATES
    for _, row in df_main.iterrows():
        processed += 1
        
        # UPDATE UI EVERY 100 ORDERS
        if processed % 100 == 0 or processed == total or processed == 1:
            progress_pct = processed / total * 100
            
            # Calculate dynamic time estimates
            if processed >= 10:  # After 10 orders, use actual performance
                elapsed = time.time() - processing_start_time
                actual_time_per_order = elapsed / processed
                remaining_orders = total - processed
                est_remaining = remaining_orders * actual_time_per_order
            else:  # For first few orders, use baseline estimate
                elapsed = time.time() - processing_start_time
                remaining_orders = total - processed
                est_remaining = remaining_orders * baseline_time_per_order
            
            if status_callback:
                strategy_name = sorting_strategy["name"] if sorting_strategy else "Default"
                if processed == total:
                    status_callback(f"✅ [Scenario {scenario_num}/{total_scenarios}] {os.path.basename(filepath)} ({strategy_name}) - Completed {total:,} orders in {elapsed:.1f}s")
                else:
                    # Show current scenario progress + context about remaining scenarios
                    remaining_scenarios = total_scenarios - scenario_num
                    if remaining_scenarios > 0:
                        status_callback(f"🔁 [Scenario {scenario_num}/{total_scenarios}] {strategy_name} - {processed:,}/{total:,} ({progress_pct:.1f}%) | {est_remaining:.0f}s + {remaining_scenarios} more")
                    else:
                        status_callback(f"🔁 [Scenario {scenario_num}/{total_scenarios}] {strategy_name} - {processed:,}/{total:,} ({progress_pct:.1f}%) | {est_remaining:.0f}s remaining")
        
        so = str(row["SO Number"]) if pd.notna(row["SO Number"]) else f"ORDER_{processed}"
        part = str(row["Part"]) if pd.notna(row["Part"]) else None
        demand_qty = row["Demand"] if pd.notna(row["Demand"]) and row["Demand"] > 0 else 0
        planner = str(row["Planner"]) if pd.notna(row["Planner"]) else "UNKNOWN"
        start_date = row["Start Date"]
        
        # Skip orders with missing critical data
        if part is None or part == "nan" or demand_qty <= 0:
            results.append({
                "SO Number": so,
                "Part": part or "MISSING",
                "Planner": planner,
                "Start Date": start_date.strftime('%Y-%m-%d') if pd.notna(start_date) else "No Date",
                "PB": "-",
                "Demand": demand_qty,
                "Hours": 0,
                "Status": "⚠️ Skipped",
                "Shortages": "-",
                "Components": "Missing part number or zero demand"
            })
            continue
        
        # Check if this is a piggyback order
        try:
            pb_check = f"NS{part}99"
            is_pb = "PB" if pb_check in planned_demand["Component Part Number"].values else "-"
        except:
            is_pb = "-"
        
        # Get planned demand for this SO
        try:
            bom = planned_demand[planned_demand["SO Number"] == so]
        except:
            bom = pd.DataFrame()
        
        # Check material availability
        releasable = True
        shortage_details = []
        components_needed = {}
        
        # Calculate labor hours for this order
        base_hours = labor_standards.get(part, 0)
        labor_hours = base_hours * demand_qty
        
        if len(bom) > 0:
            # This SO has planned component demand - use ALL-OR-NOTHING allocation
            all_components_available = True
            component_requirements = []
            
            if DEBUG_MODE and so == "9678487":
                print(f"\n=== DEBUG: Processing SO 9678487 ===")
                print(f"Found {len(bom)} components in planned demand")
            
            for _, comp in bom.iterrows():
                try:
                    comp_part = str(comp["Component Part Number"])
                    required_qty = int(comp["Component Qty Required"]) if pd.notna(comp["Component Qty Required"]) else 0
                    available = stock.get(comp_part, 0)
                    
                    # DEBUG FOR SPECIFIC PARTS
                    if DEBUG_MODE and (comp_part == "585711" or so == "9678487"):
                        print(f"\n=== DEBUG: SO {so} processing {comp_part} ===")
                        print(f"Required qty: {required_qty}")
                        print(f"Available in stock: {available}")
                        print(f"Will be sufficient: {available >= required_qty}")
                    
                    component_requirements.append({
                        'part': comp_part,
                        'required': required_qty,
                        'available': available
                    })
                    
                    components_needed[comp_part] = required_qty
                    
                    # Check availability but DON'T allocate yet
                    if available < required_qty:
                        all_components_available = False
                        shortage = required_qty - available
                        
                        if DEBUG_MODE and comp_part == "585711":
                            print(f"*** SHORTAGE DETECTED: {comp_part} short {shortage} ***")
                        
                        # Search POs for potential resolution
                        future_pos = df_pos[
                            (df_pos["Part Number"].astype(str) == comp_part) &
                            (pd.to_datetime(df_pos["Promised Due Date"], errors="coerce") >= datetime.now())
                        ]

                        po_match = None
                        for _, po_row in future_pos.iterrows():
                            po_qty = po_row["Qty Due"]
                            if po_qty >= shortage:
                                po_match = po_row
                                break

                        if po_match is not None:
                            po_id = po_match["PO Number"]
                            po_date = pd.to_datetime(po_match["Promised Due Date"]).strftime('%Y-%m-%d')
                            shortage_details.append(f"{comp_part} short {shortage} – PO {po_id} due {po_date}")
                        else:
                            shortage_details.append(f"{comp_part} (need {required_qty}, have {available}, short {shortage})")
                                
                except Exception as e:
                    all_components_available = False
                    shortage_details.append(f"Component processing error: {str(e)}")
                    continue
            
            # If ALL components available, THEN allocate all of them
            if all_components_available:
                releasable = True
                for req in component_requirements:
                    comp_part = req['part']
                    required_qty = req['required']
                    stock[comp_part] -= required_qty
            else:
                releasable = False

        else:
            # This SO has no planned component demand - treat as raw material/purchased part
            try:
                available = stock.get(part, 0)
                if available >= demand_qty:
                    stock[part] -= demand_qty
                    releasable = True
                else:
                    releasable = False
                    shortage = demand_qty - available
                    shortage_details.append(f"{part} (need {demand_qty}, have {available}, short {shortage})")
            except:
                releasable = False
                shortage_details.append(f"{part} (stock lookup failed)")

        # Build result record
        shortage_parts_only = []
        components_info = "; ".join(shortage_details) if shortage_details else str(components_needed) if components_needed else "-"

        # Extract just the part numbers from shortage details
        for detail in shortage_details:
            if " short " in detail and "–" in detail:
                part_short = detail.split(" short ")[0].strip()
                shortage_parts_only.append(part_short)
            elif "(" in detail and " (need " in detail:
                part_short = detail.split(" (need ")[0].strip()
                shortage_parts_only.append(part_short)
            elif "(" in detail:
                part_short = detail.split("(")[0].strip()
                shortage_parts_only.append(part_short)
            else:
                part_short = detail.split()[0] if detail.split() else detail
                shortage_parts_only.append(part_short)

        clean_shortages = "; ".join(shortage_parts_only) if shortage_parts_only else "-"

        results.append({
            "SO Number": so,
            "Part": part,
            "Planner": planner,
            "Start Date": start_date.strftime('%Y-%m-%d') if pd.notna(start_date) else "No Date",
            "PB": is_pb,
            "Demand": demand_qty,
            "Hours": round(labor_hours, 4),
            "Status": "✅ Release" if releasable else "❌ Hold",
            "Shortages": clean_shortages,
            "Components": components_info
        })
    
    # Calculate summary metrics
    df_results = pd.DataFrame(results)
    total_orders = len(df_results)
    releasable_count = len(df_results[df_results['Status'] == '✅ Release'])
    held_count = total_orders - releasable_count
    pb_count = len(df_results[df_results['PB'] == 'PB'])
    skipped_count = len(df_results[df_results['Status'] == '⚠️ Skipped'])
    
    total_hours = df_results['Hours'].sum()
    releasable_hours = df_results[df_results['Status'] == '✅ Release']['Hours'].sum()
    held_hours = df_results[df_results['Status'] == '❌ Hold']['Hours'].sum()
    
    # Calculate quantity metrics
    total_qty = df_results['Demand'].sum()
    releasable_qty = df_results[df_results['Status'] == '✅ Release']['Demand'].sum()
    held_qty = df_results[df_results['Status'] == '❌ Hold']['Demand'].sum()
    
   # Calculate Kit and Instrument metrics with subcategories
    releasable_results = df_results[df_results['Status'] == '✅ Release']
    
    # BVI Kits (Planner codes 3001, 3801)
    bvi_kit_planners = ['3001', '3801']
    releasable_bvi_kits = releasable_results[releasable_results['Planner'].isin(bvi_kit_planners)]
    releasable_bvi_kits_count = len(releasable_bvi_kits)
    releasable_bvi_kits_hours = releasable_bvi_kits['Hours'].sum()
    releasable_bvi_kits_qty = releasable_bvi_kits['Demand'].sum()
    
    # Malosa Kits (Planner code 5001)
    malosa_kit_planners = ['5001']
    releasable_malosa_kits = releasable_results[releasable_results['Planner'].isin(malosa_kit_planners)]
    releasable_malosa_kits_count = len(releasable_malosa_kits)
    releasable_malosa_kits_hours = releasable_malosa_kits['Hours'].sum()
    releasable_malosa_kits_qty = releasable_malosa_kits['Demand'].sum()
    
    # Total Kits (for backward compatibility)
    releasable_kits_count = releasable_bvi_kits_count + releasable_malosa_kits_count
    releasable_kits_hours = releasable_bvi_kits_hours + releasable_malosa_kits_hours
    releasable_kits_qty = releasable_bvi_kits_qty + releasable_malosa_kits_qty
    
    # Manufacturing (Planner code 3802)
    manufacturing_planners = ['3802']
    releasable_manufacturing = releasable_results[releasable_results['Planner'].isin(manufacturing_planners)]
    releasable_manufacturing_count = len(releasable_manufacturing)
    releasable_manufacturing_hours = releasable_manufacturing['Hours'].sum()
    releasable_manufacturing_qty = releasable_manufacturing['Demand'].sum()
    
    # Assembly (Planner code 3803)
    assembly_planners = ['3803']
    releasable_assembly = releasable_results[releasable_results['Planner'].isin(assembly_planners)]
    releasable_assembly_count = len(releasable_assembly)
    releasable_assembly_hours = releasable_assembly['Hours'].sum()
    releasable_assembly_qty = releasable_assembly['Demand'].sum()
    
    # Packaging (Planner code 3804)
    packaging_planners = ['3804']
    releasable_packaging = releasable_results[releasable_results['Planner'].isin(packaging_planners)]
    releasable_packaging_count = len(releasable_packaging)
    releasable_packaging_hours = releasable_packaging['Hours'].sum()
    releasable_packaging_qty = releasable_packaging['Demand'].sum()
    
    # Malosa Instruments (Planner code 3805)
    malosa_instrument_planners = ['3805']
    releasable_malosa_instruments = releasable_results[releasable_results['Planner'].isin(malosa_instrument_planners)]
    releasable_malosa_instruments_count = len(releasable_malosa_instruments)
    releasable_malosa_instruments_hours = releasable_malosa_instruments['Hours'].sum()
    releasable_malosa_instruments_qty = releasable_malosa_instruments['Demand'].sum()
    
    # Virtuoso (Planner code 3806)
    virtuoso_planners = ['3806']
    releasable_virtuoso = releasable_results[releasable_results['Planner'].isin(virtuoso_planners)]
    releasable_virtuoso_count = len(releasable_virtuoso)
    releasable_virtuoso_hours = releasable_virtuoso['Hours'].sum()
    releasable_virtuoso_qty = releasable_virtuoso['Demand'].sum()
    
    # Total Instruments (sum of all instrument categories)
    releasable_instruments_count = (releasable_manufacturing_count + releasable_assembly_count + 
                                  releasable_packaging_count + releasable_malosa_instruments_count + 
                                  releasable_virtuoso_count)
    releasable_instruments_hours = (releasable_manufacturing_hours + releasable_assembly_hours + 
                                  releasable_packaging_hours + releasable_malosa_instruments_hours + 
                                  releasable_virtuoso_hours)
    releasable_instruments_qty = (releasable_manufacturing_qty + releasable_assembly_qty + 
                                releasable_packaging_qty + releasable_malosa_instruments_qty + 
                                releasable_virtuoso_qty)

    return {
        'name': scenario_name,
        'filepath': filepath,
        'sorting_strategy': sorting_strategy["name"] if sorting_strategy else "Default (Start Date)",
        'results_df': df_results,
        'metrics': {
            'total_orders': total_orders,
            'releasable_count': releasable_count,
            'held_count': held_count,
            'pb_count': pb_count,
            'skipped_count': skipped_count,
            '---1': '---',
            'total_hours': total_hours,
            'releasable_hours': releasable_hours,
            'held_hours': held_hours,
            'total_qty': total_qty,
            'releasable_qty': releasable_qty,
            'held_qty': held_qty,
            '---2': '---',
            'releasable_kits_count': releasable_kits_count,
            'releasable_kits_hours': releasable_kits_hours,
            'releasable_kits_qty': releasable_kits_qty,
            'releasable_bvi_kits_count': releasable_bvi_kits_count,
            'releasable_bvi_kits_hours': releasable_bvi_kits_hours,
            'releasable_bvi_kits_qty': releasable_bvi_kits_qty,
            'releasable_malosa_kits_count': releasable_malosa_kits_count,
            'releasable_malosa_kits_hours': releasable_malosa_kits_hours,
            'releasable_malosa_kits_qty': releasable_malosa_kits_qty,
            '---3': '---',
            'releasable_instruments_count': releasable_instruments_count,
            'releasable_instruments_hours': releasable_instruments_hours,
            'releasable_instruments_qty': releasable_instruments_qty,
            'releasable_manufacturing_count': releasable_manufacturing_count,
            'releasable_manufacturing_hours': releasable_manufacturing_hours,
            'releasable_manufacturing_qty': releasable_manufacturing_qty,
            'releasable_assembly_count': releasable_assembly_count,
            'releasable_assembly_hours': releasable_assembly_hours,
            'releasable_assembly_qty': releasable_assembly_qty,
            'releasable_packaging_count': releasable_packaging_count,
            'releasable_packaging_hours': releasable_packaging_hours,
            'releasable_packaging_qty': releasable_packaging_qty,
            'releasable_malosa_instruments_count': releasable_malosa_instruments_count,
            'releasable_malosa_instruments_hours': releasable_malosa_instruments_hours,
            'releasable_malosa_instruments_qty': releasable_malosa_instruments_qty,
            'releasable_virtuoso_count': releasable_virtuoso_count,
            'releasable_virtuoso_hours': releasable_virtuoso_hours,
            'releasable_virtuoso_qty': releasable_virtuoso_qty,
            '---4': '---',
            'committed_parts_count': committed_parts_count,
            'total_committed_qty': total_committed_qty
        }
    }

def load_and_process_files():
    # Multiple file selection
    filepaths = filedialog.askopenfilenames(
        title="Select Material Demand Files (Hold Ctrl for multiple files)",
        filetypes=[("Excel files", "*.xlsm *.xlsx")]
    )
    if not filepaths:
        # Reset to ready state if cancelled
        status_var.set("🔄 Ready - Select Excel file(s) to begin processing")
        return

    try:
        # Clear the UI and start fresh
        status_var.set("🔄 Initializing...")
        root.update_idletasks()
        
        start_time = time.time()
        main_frame.configure(style='TFrame')
        
        scenarios = []
        scenarios_for_comparison = []  # Will store all tested scenarios for comparison tables
        
        # Progress callback function that updates UI immediately
        def update_progress(message):
            status_var.set(message)
            root.update_idletasks()
        
        # Determine processing mode
        minmax_mode = minmax_var.get()
        
        if minmax_mode:
            # Min/Max optimization mode - test all sorting strategies
            strategies = get_sorting_strategies()
            total_scenarios = len(filepaths) * len(strategies)
            
            update_progress(f"🔥 MIN/MAX MODE: Testing {len(strategies)} sorting strategies on {len(filepaths)} file(s) = {total_scenarios} total scenarios")
            time.sleep(1)
            
            scenario_num = 0
            all_strategy_results = []  # Store ALL results for comparison
            
            for file_idx, filepath in enumerate(filepaths):
                filename = os.path.basename(filepath)
                base_filename = filename.replace('.xlsm', '').replace('.xlsx', '')
                file_strategy_results = []  # Results for this specific file
                
                for strategy_idx, strategy in enumerate(strategies):
                    scenario_num += 1
                    scenario_name = f"{base_filename}_{strategy['name'].replace(' ', '_').replace('(', '').replace(')', '')}"
                    
                    # Process with specific sorting strategy
                    scenario_start_time = time.time()
                    scenario_result = process_single_scenario(
                        filepath, scenario_name, update_progress, 
                        scenario_num, total_scenarios, strategy
                    )
                    scenario_end_time = time.time()
                    scenario_duration = scenario_end_time - scenario_start_time
                    
                    # Store result for this file
                    file_strategy_results.append(scenario_result)
                    all_strategy_results.append(scenario_result)
                    
                    # Show completion
                    metrics = scenario_result['metrics']
                    remaining_scenarios = total_scenarios - scenario_num
                    
                    if remaining_scenarios > 0:
                        total_elapsed = time.time() - start_time
                        avg_time_per_scenario = total_elapsed / scenario_num
                        estimated_remaining = remaining_scenarios * avg_time_per_scenario
                        
                        update_progress(f"✅ [{scenario_num}/{total_scenarios}] {strategy['name']}: {metrics['releasable_count']:,}/{metrics['total_orders']:,} orders ({scenario_duration:.1f}s) | {estimated_remaining:.0f}s remaining")
                    else:
                        update_progress(f"✅ [{scenario_num}/{total_scenarios}] {strategy['name']}: {metrics['releasable_count']:,}/{metrics['total_orders']:,} orders ({scenario_duration:.1f}s) | OPTIMIZATION COMPLETE!")
                    
                    time.sleep(0.2)  # Brief pause between strategies
                
                # After testing all strategies for this file, find the best ones
                best_orders_strategy = max(file_strategy_results, key=lambda s: s['metrics']['releasable_count'])
                best_hours_strategy = max(file_strategy_results, key=lambda s: s['metrics']['releasable_hours'])
                best_qty_strategy = max(file_strategy_results, key=lambda s: s['metrics']['releasable_qty'])
                
                # Create NEW scenario objects with clear names for the best strategies
                # Best Orders Strategy
                best_orders_scenario = {
                    'name': f"BEST_ORDERS_{base_filename}",
                    'filepath': filepath,
                    'sorting_strategy': f"🏆 BEST ORDERS: {best_orders_strategy['sorting_strategy']}",
                    'results_df': best_orders_strategy['results_df'],
                    'metrics': best_orders_strategy['metrics']
                }
                scenarios.append(best_orders_scenario)
                
                # Best Hours Strategy
                best_hours_scenario = {
                    'name': f"BEST_HOURS_{base_filename}",
                    'filepath': filepath,
                    'sorting_strategy': f"🏆 BEST HOURS: {best_hours_strategy['sorting_strategy']}",
                    'results_df': best_hours_strategy['results_df'],
                    'metrics': best_hours_strategy['metrics']
                }
                scenarios.append(best_hours_scenario)
                
                # Best Quantity Strategy
                best_qty_scenario = {
                    'name': f"BEST_QTY_{base_filename}",
                    'filepath': filepath,
                    'sorting_strategy': f"🏆 BEST QTY: {best_qty_strategy['sorting_strategy']}",
                    'results_df': best_qty_strategy['results_df'],
                    'metrics': best_qty_strategy['metrics']
                }
                scenarios.append(best_qty_scenario)
                
                update_progress(f"🏆 File {file_idx+1}/{len(filepaths)} optimized: Orders={best_orders_strategy['sorting_strategy']} ({best_orders_strategy['metrics']['releasable_count']:,}), Hours={best_hours_strategy['sorting_strategy']} ({best_hours_strategy['metrics']['releasable_hours']:,.0f}), Qty={best_qty_strategy['sorting_strategy']} ({best_qty_strategy['metrics']['releasable_qty']:,})")
                time.sleep(0.5)
            
            # Use all_strategy_results for comparison tables
            scenarios_for_comparison = all_strategy_results
        else:
            # Standard mode - process files normally
            total_scenarios = len(filepaths)
            
            for i, filepath in enumerate(filepaths):
                filename = os.path.basename(filepath)
                scenario_name = f"Scenario_{i+1}_{filename.replace('.xlsm', '').replace('.xlsx', '')}"
                scenario_num = i + 1
                
                # Calculate estimate for this scenario
                if i > 0:
                    total_elapsed = time.time() - start_time
                    avg_time_per_scenario = total_elapsed / i
                    remaining_scenarios = len(filepaths) - i
                    total_est_remaining = remaining_scenarios * avg_time_per_scenario
                    update_progress(f"📊 [Scenario {scenario_num}/{len(filepaths)}] Starting: {filename} | Est. {total_est_remaining:.0f}s for all remaining scenarios")
                else:
                    update_progress(f"📊 [Scenario {scenario_num}/{len(filepaths)}] Starting: {filename}")
                
                # Process with live progress updates
                scenario_start_time = time.time()
                scenario_result = process_single_scenario(filepath, scenario_name, update_progress, scenario_num, len(filepaths))
                scenario_end_time = time.time()
                scenario_duration = scenario_end_time - scenario_start_time
                
                scenarios.append(scenario_result)
                scenarios_for_comparison.append(scenario_result)  # Same as scenarios in standard mode
                
                # Show completion with actual metrics and time
                metrics = scenario_result['metrics']
                
                if i < len(filepaths) - 1:
                    total_elapsed = time.time() - start_time
                    avg_time_per_scenario = total_elapsed / (i + 1)
                    remaining_scenarios = len(filepaths) - (i + 1)
                    estimated_remaining = remaining_scenarios * avg_time_per_scenario
                    
                    update_progress(f"✅ [Scenario {scenario_num}/{len(filepaths)}] Complete: {metrics['releasable_count']:,}/{metrics['total_orders']:,} releasable ({scenario_duration:.1f}s) | Est. {estimated_remaining:.0f}s for {remaining_scenarios} remaining scenarios")
                else:
                    update_progress(f"✅ [Scenario {scenario_num}/{len(filepaths)}] Complete: {metrics['releasable_count']:,}/{metrics['total_orders']:,} releasable ({scenario_duration:.1f}s) | COMPLETE!")
                
                time.sleep(0.3)
        
        # Calculate total processing time
        end_time = time.time()
        processing_time = end_time - start_time
        
        # Helper function to aggregate metrics from a list of scenarios
        def aggregate_metrics(scenario_list, is_min_max_mode_for_totals=False):
            if not scenario_list:
                # Return a default structure if the list is empty
                return {key: 0 for key in [
                    'total_qty', 'total_hours',
                    'releasable_count', 'releasable_qty', 'releasable_hours',
                    'releasable_kits_count', 'releasable_kits_hours', 'releasable_kits_qty',
                    'releasable_bvi_kits_count', 'releasable_bvi_kits_hours', 'releasable_bvi_kits_qty',
                    'releasable_malosa_kits_count', 'releasable_malosa_kits_hours', 'releasable_malosa_kits_qty',
                    'releasable_instruments_count', 'releasable_instruments_hours', 'releasable_instruments_qty',
                    'releasable_manufacturing_count', 'releasable_manufacturing_hours', 'releasable_manufacturing_qty',
                    'releasable_assembly_count', 'releasable_assembly_hours', 'releasable_assembly_qty',
                    'releasable_packaging_count', 'releasable_packaging_hours', 'releasable_packaging_qty',
                    'releasable_malosa_instruments_count', 'releasable_malosa_instruments_hours', 'releasable_malosa_instruments_qty',
                    'releasable_virtuoso_count', 'releasable_virtuoso_hours', 'releasable_virtuoso_qty'
                ]}

            agg = {}
            keys_to_sum = [
                'releasable_count', 'releasable_qty', 'releasable_hours',
                'releasable_kits_count', 'releasable_kits_hours', 'releasable_kits_qty',
                'releasable_bvi_kits_count', 'releasable_bvi_kits_hours', 'releasable_bvi_kits_qty',
                'releasable_malosa_kits_count', 'releasable_malosa_kits_hours', 'releasable_malosa_kits_qty',
                'releasable_instruments_count', 'releasable_instruments_hours', 'releasable_instruments_qty',
                'releasable_manufacturing_count', 'releasable_manufacturing_hours', 'releasable_manufacturing_qty',
                'releasable_assembly_count', 'releasable_assembly_hours', 'releasable_assembly_qty',
                'releasable_packaging_count', 'releasable_packaging_hours', 'releasable_packaging_qty',
                'releasable_malosa_instruments_count', 'releasable_malosa_instruments_hours', 'releasable_malosa_instruments_qty',
                'releasable_virtuoso_count', 'releasable_virtuoso_hours', 'releasable_virtuoso_qty',
                'total_qty', 'total_hours' # These are per-scenario totals, summing them gives overall totals
            ]
            for key in keys_to_sum:
                agg[key] = sum(s['metrics'].get(key, 0) for s in scenario_list)
            
            return agg

        # Determine source metrics for the Summary sheet
        source_metrics = {}
        total_orders_processed = 0

        if not scenarios: # Handle case where no scenarios were processed (e.g., file dialog cancelled after initial selection)
            # Use default empty metrics for the summary if no scenarios exist
            source_metrics = aggregate_metrics([], is_min_max_mode_for_totals=minmax_mode)
            # total_orders_processed is already 0
            # Ensure filepaths is not empty if we try to use it later for "Files Processed"
            files_processed_list_for_summary = []
        else:
            files_processed_list_for_summary = filepaths # Use the original filepaths list for the summary

            if minmax_mode:
                # In min/max mode, total_orders_processed sums total_orders from one "best" strategy per unique file
                seen_files_for_total_orders = set()
                unique_scenarios_for_total_orders = []
                for s in scenarios: # 'scenarios' contains BEST_ORDERS, BEST_HOURS, BEST_QTY per file
                    if s['filepath'] not in seen_files_for_total_orders:
                        unique_scenarios_for_total_orders.append(s)
                        seen_files_for_total_orders.add(s['filepath'])
                total_orders_processed = sum(s['metrics']['total_orders'] for s in unique_scenarios_for_total_orders)
                
                # For detailed metrics in summary, aggregate from "BEST_ORDERS" strategies
                best_orders_scenarios = [s for s in scenarios if "BEST_ORDERS" in s['name']]
                source_metrics = aggregate_metrics(best_orders_scenarios, is_min_max_mode_for_totals=True)
            else: # Standard mode
                total_orders_processed = sum(s['metrics']['total_orders'] for s in scenarios)
                if len(scenarios) == 1:
                    source_metrics = scenarios[0]['metrics']
                else: # Standard mode, multiple files
                    source_metrics = aggregate_metrics(scenarios)
        
        orders_per_second = total_orders_processed / processing_time if processing_time > 0 else 0
        
        # Save results
        if not no_export_var.get():
            # Save results
            output_dir = os.path.dirname(filepaths[0])
            timestamp = datetime.now().strftime('%Y%m%d_%H%M')
            
            if minmax_mode:
                output_file = os.path.join(output_dir, f"MinMax_Optimization_Analysis_{VERSION}_{timestamp}.xlsx")
            elif len(scenarios) > 1:
                output_file = os.path.join(output_dir, f"Multi_Scenario_Analysis_{VERSION}_{timestamp}.xlsx")
            else:
                output_file = os.path.join(output_dir, f"Material_Release_Plan_{VERSION}_{timestamp}.xlsx")
            
            status_var.set("💾 Saving optimization results...")
            root.update_idletasks()
            
            # Create summary sheet
            summary_items = [
                ('Tool Version', f"{VERSION} ({VERSION_DATE})"),
                ('Processing Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
                ('Processing Mode', "Min/Max Optimization" if minmax_mode else "Standard"),
                ('Files Processed Count', len(files_processed_list_for_summary)),
                ('Strategies Tested / Scenarios', len(scenarios_for_comparison) if minmax_mode else len(scenarios)),
                ('Optimal Strategies / Scenarios Saved', len(scenarios)), # In minmax, 'scenarios' has BEST_xxx. In standard, it's all scenarios.
                ('--- Overall Performance ---', '---'),
                ('Total Orders Processed', f"{total_orders_processed:,}"),
                ('Total Demand Quantity', f"{source_metrics.get('total_qty', 0):,}"),
                ('Total Demand Hours', f"{source_metrics.get('total_hours', 0):,.1f}"),
                ('Releasable Orders', f"{source_metrics.get('releasable_count', 0):,}"),
                ('Releasable Quantity', f"{source_metrics.get('releasable_qty', 0):,.1f}"),
                ('Releasable Hours', f"{source_metrics.get('releasable_hours', 0):,.1f}"),
                ('--- Releasable Kits Breakdown ---', '---'),
                ('Total Releasable Kits (Orders)', f"{source_metrics.get('releasable_kits_count', 0):,}"),
                ('Total Releasable Kits (Hours)', f"{source_metrics.get('releasable_kits_hours', 0):,.1f}"),
                ('Total Releasable Kits (Quantity)', f"{source_metrics.get('releasable_kits_qty', 0):,}"),
                ('  BVI Kits (Orders)', f"{source_metrics.get('releasable_bvi_kits_count', 0):,}"),
                ('  BVI Kits (Hours)', f"{source_metrics.get('releasable_bvi_kits_hours', 0):,.1f}"),
                ('  BVI Kits (Quantity)', f"{source_metrics.get('releasable_bvi_kits_qty', 0):,}"),
                ('  Malosa Kits (Orders)', f"{source_metrics.get('releasable_malosa_kits_count', 0):,}"),
                ('  Malosa Kits (Hours)', f"{source_metrics.get('releasable_malosa_kits_hours', 0):,.1f}"),
                ('  Malosa Kits (Quantity)', f"{source_metrics.get('releasable_malosa_kits_qty', 0):,}"),
                ('--- Releasable Instruments Breakdown ---', '---'),
                ('Total Releasable Instruments (Orders)', f"{source_metrics.get('releasable_instruments_count', 0):,}"),
                ('Total Releasable Instruments (Hours)', f"{source_metrics.get('releasable_instruments_hours', 0):,.1f}"),
                ('Total Releasable Instruments (Quantity)', f"{source_metrics.get('releasable_instruments_qty', 0):,}"),
                ('  Manufacturing (3802 Orders)', f"{source_metrics.get('releasable_manufacturing_count', 0):,}"),
                ('  Manufacturing (3802 Hours)', f"{source_metrics.get('releasable_manufacturing_hours', 0):,.1f}"),
                ('  Manufacturing (3802 Quantity)', f"{source_metrics.get('releasable_manufacturing_qty', 0):,}"),
                ('  Assembly (3803 Orders)', f"{source_metrics.get('releasable_assembly_count', 0):,}"),
                ('  Assembly (3803 Hours)', f"{source_metrics.get('releasable_assembly_hours', 0):,.1f}"),
                ('  Assembly (3803 Quantity)', f"{source_metrics.get('releasable_assembly_qty', 0):,}"),
                ('  Packaging (3804 Orders)', f"{source_metrics.get('releasable_packaging_count', 0):,}"),
                ('  Packaging (3804 Hours)', f"{source_metrics.get('releasable_packaging_hours', 0):,.1f}"),
                ('  Packaging (3804 Quantity)', f"{source_metrics.get('releasable_packaging_qty', 0):,}"),
                ('  Malosa Instruments (3805 Orders)', f"{source_metrics.get('releasable_malosa_instruments_count', 0):,}"),
                ('  Malosa Instruments (3805 Hours)', f"{source_metrics.get('releasable_malosa_instruments_hours', 0):,.1f}"),
                ('  Malosa Instruments (3805 Quantity)', f"{source_metrics.get('releasable_malosa_instruments_qty', 0):,}"),
                ('  Virtuoso (3806 Orders)', f"{source_metrics.get('releasable_virtuoso_count', 0):,}"),
                ('  Virtuoso (3806 Hours)', f"{source_metrics.get('releasable_virtuoso_hours', 0):,.1f}"),
                ('  Virtuoso (3806 Quantity)', f"{source_metrics.get('releasable_virtuoso_qty', 0):,}"),
                ('--- Processing Performance ---', '---'),
                ('Total Processing Time (seconds)', f"{processing_time:.2f}"),
                ('Processing Speed (orders/second)', f"{orders_per_second:.1f}"),
                ('Avg Time per Strategy/Scenario (seconds)', f"{processing_time/len(scenarios_for_comparison):.2f}" if minmax_mode and scenarios_for_comparison else (f"{processing_time/len(scenarios):.2f}" if scenarios else "N/A")),
                ('Processed File Names', "; ".join([os.path.basename(f) for f in files_processed_list_for_summary]))
            ]
            summary_data = pd.DataFrame(summary_items, columns=['Metric', 'Value'])
            
            # Create comparison data if multiple scenarios
            if len(scenarios_for_comparison) > 1:
                comparison_data = []
                for scenario in scenarios_for_comparison:
                    metrics = scenario['metrics']
                    comparison_data.append({
                        'Scenario': scenario['name'],
                        'File': os.path.basename(scenario['filepath']),
                        'Sorting Strategy': scenario['sorting_strategy'],
                        'Total Orders': metrics['total_orders'],
                        'Releasable Orders': metrics['releasable_count'],
                        'Held Orders': metrics['held_count'],
                        'Release Rate (%)': f"{metrics['releasable_count']/metrics['total_orders']*100:.1f}%" if metrics['total_orders'] > 0 else "0%",
                        '---1': '---',
                        'Total Qty': f"{metrics['total_qty']:,}",
                        'Releasable Qty': f"{metrics['releasable_qty']:,}",
                        'Qty Release Rate (%)': f"{metrics['releasable_qty']/metrics['total_qty']*100:.1f}%" if metrics['total_qty'] > 0 else "0%",
                        'Piggyback Orders': metrics['pb_count'],
                        'Total Hours': f"{metrics['total_hours']:,.1f}",
                        'Releasable Hours': f"{metrics['releasable_hours']:,.1f}",
                        'Labor Release Rate (%)': f"{metrics['releasable_hours']/metrics['total_hours']*100:.1f}%" if metrics['total_hours'] > 0 else "0%",
                        '---2': '---',
                        'BVI Kits': metrics['releasable_bvi_kits_count'],
                        'BVI Kit Hours': f"{metrics['releasable_bvi_kits_hours']:,.1f}",
                        'BVI Kit Qty': f"{metrics['releasable_bvi_kits_qty']:,}",
                        'Malosa Kits': metrics['releasable_malosa_kits_count'],
                        'Malosa Kit Hours': f"{metrics['releasable_malosa_kits_hours']:,.1f}",
                        'Malosa Kit Qty': f"{metrics['releasable_malosa_kits_qty']:,}",
                        'Total Kits': metrics['releasable_kits_count'],
                        'Total Kit Hours': f"{metrics['releasable_kits_hours']:,.1f}",
                        'Total Kit Qty': f"{metrics['releasable_kits_qty']:,}",
                        '---3': '---',
                        'Manufacturing (3802)': metrics['releasable_manufacturing_count'],
                        'Manufacturing Hours': f"{metrics['releasable_manufacturing_hours']:,.1f}",
                        'Manufacturing Qty': f"{metrics['releasable_manufacturing_qty']:,}",
                        'Assembly (3803)': metrics['releasable_assembly_count'],
                        'Assembly Hours': f"{metrics['releasable_assembly_hours']:,.1f}",
                        'Assembly Qty': f"{metrics['releasable_assembly_qty']:,}",
                        'Packaging (3804)': metrics['releasable_packaging_count'],
                        'Packaging Hours': f"{metrics['releasable_packaging_hours']:,.1f}",
                        'Packaging Qty': f"{metrics['releasable_packaging_qty']:,}",
                        'Malosa Inst (3805)': metrics['releasable_malosa_instruments_count'],
                        'Malosa Inst Hours': f"{metrics['releasable_malosa_instruments_hours']:,.1f}",
                        'Malosa Inst Qty': f"{metrics['releasable_malosa_instruments_qty']:,}",
                        'Virtuoso (3806)': metrics['releasable_virtuoso_count'],
                        'Virtuoso Hours': f"{metrics['releasable_virtuoso_hours']:,.1f}",
                        'Virtuoso Qty': f"{metrics['releasable_virtuoso_qty']:,}",
                        'Total Instruments': metrics['releasable_instruments_count'],
                        'Total Inst Hours': f"{metrics['releasable_instruments_hours']:,.1f}",
                        'Total Inst Qty': f"{metrics['releasable_instruments_qty']:,}",
                        '---4': '---',
                        'Committed Parts': metrics['committed_parts_count'],
                        'Committed Qty': f"{metrics['total_committed_qty']:,}"
                    })
                comparison_df = pd.DataFrame(comparison_data)

            # Write everything to Excel in a single writer session
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # Write each scenario to its own sheet first
                for scenario in scenarios:
                    sheet_name = scenario['name'][:31]  # Excel sheet name limit
                    scenario['results_df'].to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Write the summary sheet with formatting
                summary_data.to_excel(writer, sheet_name='Summary', index=False)
                
                # Write comparison sheet if it exists
                if len(scenarios_for_comparison) > 1:
                    comparison_df.to_excel(writer, sheet_name='Strategy Comparison', index=False)
                
                # Get the workbook
                workbook = writer.book
                
                # Format Summary sheet
                worksheet = writer.sheets['Summary']
                
                # Define styles
                header_fill = openpyxl.styles.PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
                separator_fill = openpyxl.styles.PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
                bold_font = openpyxl.styles.Font(bold=True)
                
                # Apply formatting to Summary sheet
                for row in range(2, worksheet.max_row + 1):  # Start from row 2 (after header)
                    cell_value = worksheet.cell(row=row, column=1).value
                    if cell_value.startswith('---'):
                        # Apply separator formatting
                        for col in range(1, 3):  # Columns A and B
                            cell = worksheet.cell(row=row, column=col)
                            cell.fill = separator_fill
                    elif any(cell_value.startswith(prefix) for prefix in ['Total', 'Releasable', 'BVI', 'Malosa', 'Manufacturing', 'Assembly', 'Packaging', 'Virtuoso']):
                        # Bold important metrics
                        worksheet.cell(row=row, column=1).font = bold_font
                
                # Format header row in Summary
                for col in range(1, 3):  # Columns A and B
                    cell = worksheet.cell(row=1, column=col)
                    cell.font = bold_font
                    cell.fill = header_fill
                
                # Auto-adjust column widths in Summary
                for column in worksheet.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                        adjusted_width = (max_length + 2)
                        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
                
                # If comparison sheet exists, format it too
                if len(scenarios_for_comparison) > 1:
                    comp_worksheet = writer.sheets['Strategy Comparison']
                    
                    # Format header row in Comparison
                    for col in range(1, comp_worksheet.max_column + 1):
                        cell = comp_worksheet.cell(row=1, column=col)
                        cell.font = bold_font
                        cell.fill = header_fill
                    
                    # Auto-adjust column widths in Comparison
                    for column in comp_worksheet.columns:
                        max_length = 0
                        column = [cell for cell in column]
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                            adjusted_width = min((max_length + 2), 50)  # Cap width at 50
                            comp_worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
        else:
            output_file = None  # No file created

        # Display results
        if minmax_mode:
            # Min/Max optimization summary
            files_processed = list(set([s['filepath'] for s in scenarios_for_comparison]))
            
            summary_text = f"""🔥 MIN/MAX OPTIMIZATION COMPLETE!

📊 OPTIMIZATION ANALYSIS:
   Files Analyzed: {len(files_processed)}
   Sorting Strategies Tested: {len(get_sorting_strategies())}
   Total Strategy Tests: {len(scenarios_for_comparison)}
   Best Strategies Saved: {len(scenarios)} individual sheets (3 per file: Orders, Hours, Qty)

"""
            
            for filepath in files_processed:
                file_scenarios = [s for s in scenarios_for_comparison if s['filepath'] == filepath]
                best_orders = max(file_scenarios, key=lambda s: s['metrics']['releasable_count'])
                best_hours = max(file_scenarios, key=lambda s: s['metrics']['releasable_hours'])
                best_qty = max(file_scenarios, key=lambda s: s['metrics']['releasable_qty'])
                worst_orders = min(file_scenarios, key=lambda s: s['metrics']['releasable_count'])
                
                improvement_orders = best_orders['metrics']['releasable_count'] - worst_orders['metrics']['releasable_count']
                improvement_pct = improvement_orders / worst_orders['metrics']['total_orders'] * 100
                
                summary_text += f"""📁 FILE: {os.path.basename(filepath)}
   🏆 BEST STRATEGY (Orders): {best_orders['sorting_strategy']}
      → {best_orders['metrics']['releasable_count']:,}/{best_orders['metrics']['total_orders']:,} orders releasable ({best_orders['metrics']['releasable_count']/best_orders['metrics']['total_orders']*100:.1f}%)
        🔧 BVI Kits: {best_orders['metrics']['releasable_bvi_kits_count']:,} orders, {best_orders['metrics']['releasable_bvi_kits_hours']:,.0f} hrs, {best_orders['metrics']['releasable_bvi_kits_qty']:,} qty
        🔧 Malosa Kits: {best_orders['metrics']['releasable_malosa_kits_count']:,} orders, {best_orders['metrics']['releasable_malosa_kits_hours']:,.0f} hrs, {best_orders['metrics']['releasable_malosa_kits_qty']:,} qty
        🔬 Manufacturing: {best_orders['metrics']['releasable_manufacturing_count']:,} orders, {best_orders['metrics']['releasable_manufacturing_hours']:,.0f} hrs, {best_orders['metrics']['releasable_manufacturing_qty']:,} qty
        🔧 Assembly: {best_orders['metrics']['releasable_assembly_count']:,} orders, {best_orders['metrics']['releasable_assembly_hours']:,.0f} hrs, {best_orders['metrics']['releasable_assembly_qty']:,} qty
        📦 Packaging: {best_orders['metrics']['releasable_packaging_count']:,} orders, {best_orders['metrics']['releasable_packaging_hours']:,.0f} hrs, {best_orders['metrics']['releasable_packaging_qty']:,} qty
        🔬 Malosa Instruments: {best_orders['metrics']['releasable_malosa_instruments_count']:,} orders, {best_orders['metrics']['releasable_malosa_instruments_hours']:,.0f} hrs, {best_orders['metrics']['releasable_malosa_instruments_qty']:,} qty
        🎵 Virtuoso: {best_orders['metrics']['releasable_virtuoso_count']:,} orders, {best_orders['metrics']['releasable_virtuoso_hours']:,.0f} hrs, {best_orders['metrics']['releasable_virtuoso_qty']:,} qty

   🏆 BEST STRATEGY (Hours): {best_hours['sorting_strategy']}
      → {best_hours['metrics']['releasable_hours']:,.0f}/{best_hours['metrics']['total_hours']:,.0f} hours releasable ({best_hours['metrics']['releasable_hours']/best_hours['metrics']['total_hours']*100:.1f}%)
   
   🏆 BEST STRATEGY (Qty): {best_qty['sorting_strategy']}
      → {best_qty['metrics']['releasable_qty']:,}/{best_qty['metrics']['total_qty']:,} units releasable ({best_qty['metrics']['releasable_qty']/best_qty['metrics']['total_qty']*100:.1f}%)
   
   📉 WORST STRATEGY: {worst_orders['sorting_strategy']}
      → {worst_orders['metrics']['releasable_count']:,} orders releasable
   
   🔺 IMPROVEMENT POTENTIAL: +{improvement_orders:,} more orders ({improvement_pct:.1f}% boost)

"""
            
            summary_text += f"""⏱️ PERFORMANCE METRICS:
   Total Processing Time: {processing_time:.2f} seconds
   Processing Speed: {orders_per_second:.1f} orders/second
   Average per Strategy: {processing_time/len(scenarios_for_comparison):.1f} seconds
   
💾 Results saved to: {os.path.basename(output_file) if output_file else 'No export (Quick Analysis Mode)'}
   
🔥 OPTIMIZATION FEATURES:
   ✓ All sorting strategies tested
   ✓ Triple optimization: Orders + Hours + Quantity
   ✓ Only optimal results saved as individual sheets
   ✓ Complete strategy comparison table
   ✓ Improvement potential analysis"""
            
        elif len(scenarios) > 1:
            # Multi-scenario summary (standard mode)
            best_scenario = max(scenarios, key=lambda s: s['metrics']['releasable_count'])
            worst_scenario = min(scenarios, key=lambda s: s['metrics']['releasable_count'])
            improvement = best_scenario['metrics']['releasable_count'] - worst_scenario['metrics']['releasable_count']
            
            summary_text = f"""✅ MULTI-SCENARIO ANALYSIS COMPLETE!

📊 SCENARIOS COMPARED: {len(scenarios)}

🏆 BEST PERFORMER: {os.path.basename(best_scenario['filepath'])}
   ✅ {best_scenario['metrics']['releasable_count']:,} releasable orders ({best_scenario['metrics']['releasable_count']/best_scenario['metrics']['total_orders']*100:.1f}%)
   🔧 BVI Kits: {best_scenario['metrics']['releasable_bvi_kits_count']:,} orders, {best_scenario['metrics']['releasable_bvi_kits_hours']:,.0f} hrs, {best_scenario['metrics']['releasable_bvi_kits_qty']:,} qty
   🔧 Malosa Kits: {best_scenario['metrics']['releasable_malosa_kits_count']:,} orders, {best_scenario['metrics']['releasable_malosa_kits_hours']:,.0f} hrs, {best_scenario['metrics']['releasable_malosa_kits_qty']:,} qty
   🔬 Manufacturing: {best_scenario['metrics']['releasable_manufacturing_count']:,} orders, {best_scenario['metrics']['releasable_manufacturing_hours']:,.0f} hrs, {best_scenario['metrics']['releasable_manufacturing_qty']:,} qty
   🔧 Assembly: {best_scenario['metrics']['releasable_assembly_count']:,} orders, {best_scenario['metrics']['releasable_assembly_hours']:,.0f} hrs, {best_scenario['metrics']['releasable_assembly_qty']:,} qty
   📦 Packaging: {best_scenario['metrics']['releasable_packaging_count']:,} orders, {best_scenario['metrics']['releasable_packaging_hours']:,.0f} hrs, {best_scenario['metrics']['releasable_packaging_qty']:,} qty
   🔬 Malosa Instruments: {best_scenario['metrics']['releasable_malosa_instruments_count']:,} orders, {best_scenario['metrics']['releasable_malosa_instruments_hours']:,.0f} hrs, {best_scenario['metrics']['releasable_malosa_instruments_qty']:,} qty
   🎵 Virtuoso: {best_scenario['metrics']['releasable_virtuoso_count']:,} orders, {best_scenario['metrics']['releasable_virtuoso_hours']:,.0f} hrs, {best_scenario['metrics']['releasable_virtuoso_qty']:,} qty

📉 BASELINE: {os.path.basename(worst_scenario['filepath'])}
   ✅ {worst_scenario['metrics']['releasable_count']:,} releasable orders ({worst_scenario['metrics']['releasable_count']/worst_scenario['metrics']['total_orders']*100:.1f}%)

🔺 IMPROVEMENT: +{improvement:,} more orders releasable

⏱️ PERFORMANCE METRICS:
   Total Processing Time: {processing_time:.2f} seconds
   Processing Speed: {orders_per_second:.1f} orders/second
   
💾 Results saved to: {os.path.basename(output_file) if output_file else 'No export (Quick Analysis Mode)'}"""
        else:
            # Single scenario summary
            scenario = scenarios[0]
            metrics = scenario['metrics']
            
            summary_text = f"""✅ PROCESSING COMPLETE!

📊 RESULTS SUMMARY:
   Total Orders: {format_metric(safe_metric(metrics, 'total_orders'))}
   ✅ Releasable: {format_metric(safe_metric(metrics, 'releasable_count'))} ({format_metric(safe_metric(metrics, 'releasable_count') / safe_metric(metrics, 'total_orders') * 100, 'percentage')})
   ❌ On Hold: {format_metric(safe_metric(metrics, 'held_count'))} ({format_metric(safe_metric(metrics, 'held_count') / safe_metric(metrics, 'total_orders') * 100, 'percentage')})
   🏷️ Piggyback: {format_metric(safe_metric(metrics, 'pb_count'))}
   ⚠️ Skipped: {format_metric(safe_metric(metrics, 'skipped_count'))}

🔧 RELEASABLE KITS:
   BVI Kits (3001, 3801): {format_metric(safe_metric(metrics, 'releasable_bvi_kits_count'))} orders, {format_metric(safe_metric(metrics, 'releasable_bvi_kits_hours'), 'hours')} hrs, {format_metric(safe_metric(metrics, 'releasable_bvi_kits_qty'))} qty
   Malosa Kits (5001): {format_metric(safe_metric(metrics, 'releasable_malosa_kits_count'))} orders, {format_metric(safe_metric(metrics, 'releasable_malosa_kits_hours'), 'hours')} hrs, {format_metric(safe_metric(metrics, 'releasable_malosa_kits_qty'))} qty
   Total Kits: {format_metric(safe_metric(metrics, 'releasable_kits_count'))} orders, {format_metric(safe_metric(metrics, 'releasable_kits_hours'), 'hours')} hrs, {format_metric(safe_metric(metrics, 'releasable_kits_qty'))} qty

🔬 RELEASABLE INSTRUMENTS:
   Manufacturing (3802): {format_metric(safe_metric(metrics, 'releasable_manufacturing_count'))} orders, {format_metric(safe_metric(metrics, 'releasable_manufacturing_hours'), 'hours')} hrs, {format_metric(safe_metric(metrics, 'releasable_manufacturing_qty'))} qty
   Assembly (3803): {format_metric(safe_metric(metrics, 'releasable_assembly_count'))} orders, {format_metric(safe_metric(metrics, 'releasable_assembly_hours'), 'hours')} hrs, {format_metric(safe_metric(metrics, 'releasable_assembly_qty'))} qty
   Packaging (3804): {format_metric(safe_metric(metrics, 'releasable_packaging_count'))} orders, {format_metric(safe_metric(metrics, 'releasable_packaging_hours'), 'hours')} hrs, {format_metric(safe_metric(metrics, 'releasable_packaging_qty'))} qty
   Malosa Instruments (3805): {format_metric(safe_metric(metrics, 'releasable_malosa_instruments_count'))} orders, {format_metric(safe_metric(metrics, 'releasable_malosa_instruments_hours'), 'hours')} hrs, {format_metric(safe_metric(metrics, 'releasable_malosa_instruments_qty'))} qty
   Virtuoso (3806): {format_metric(safe_metric(metrics, 'releasable_virtuoso_count'))} orders, {format_metric(safe_metric(metrics, 'releasable_virtuoso_hours'), 'hours')} hrs, {format_metric(safe_metric(metrics, 'releasable_virtuoso_qty'))} qty
   Total Instruments: {format_metric(safe_metric(metrics, 'releasable_instruments_count'))} orders, {format_metric(safe_metric(metrics, 'releasable_instruments_hours'), 'hours')} hrs, {format_metric(safe_metric(metrics, 'releasable_instruments_qty'))} qty

⏱️ LABOR HOURS SUMMARY:
   Total Hours: {format_metric(safe_metric(metrics, 'total_hours'), 'hours')}
   ✅ Releasable Hours: {format_metric(safe_metric(metrics, 'releasable_hours'), 'hours')} ({format_metric(safe_metric(metrics, 'releasable_hours') / safe_metric(metrics, 'total_hours') * 100, 'percentage')})

⏱️ PERFORMANCE METRICS:
   Processing Time: {processing_time:.2f} seconds
   Orders per Second: {orders_per_second:.1f}

💾 Results saved to: {os.path.basename(output_file) if output_file else 'No export (Quick Analysis Mode)'}"""
        
        results_text.delete(1.0, tk.END)
        results_text.insert(1.0, summary_text)
        
        # For status bar
        if minmax_mode:
            status_var.set(f"🔥 MIN/MAX OPTIMIZATION COMPLETE! {len(scenarios_for_comparison)} strategies tested, {len(scenarios)} best results saved in {processing_time:.1f}s")
        elif len(scenarios) > 1:
            best_scenario = max(scenarios, key=lambda s: s['metrics']['releasable_count'])
            worst_scenario = min(scenarios, key=lambda s: s['metrics']['releasable_count'])
            improvement = best_scenario['metrics']['releasable_count'] - worst_scenario['metrics']['releasable_count']
            status_var.set(f"✅ ALL {len(scenarios)} SCENARIOS COMPLETE! Best: {best_scenario['metrics']['releasable_count']:,} releasable (+{improvement:,} vs worst) | Total time: {processing_time:.1f}s")
        else:
            total_orders = scenarios[0]['metrics']['total_orders']
            total_releasable = scenarios[0]['metrics']['releasable_count']
            status_var.set(f"✅ PROCESSING COMPLETE! {total_releasable:,}/{total_orders:,} orders releasable in {processing_time:.1f}s")
            
        main_frame.configure(style='Success.TFrame')
        
    except Exception as e:
        messagebox.showerror("Error", f"Processing failed:\n\n{str(e)}")
        status_var.set("❌ Processing failed")

def copy_summary_to_clipboard():
    summary_text_content = results_text.get("1.0", tk.END).strip()
    root.clipboard_clear()
    root.clipboard_append(summary_text_content)
    root.update()
    copy_btn.config(text="✅ Copied!")
    root.after(2000, lambda: copy_btn.config(text="📋 Copy Summary"))

# Create GUI
root = tk.Tk()
root.title(f"PlanSnap {VERSION} - Material Release Planning Tool")
root.geometry("600x650")

# Main frame
main_frame = ttk.Frame(root, padding="20")
main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

# Title
title_label = ttk.Label(main_frame, text=f"PlanSnap {VERSION}", 
                       font=('Arial', 16, 'bold'))
title_label.grid(row=0, column=0, pady=(0, 5))

# Version date
version_label = ttk.Label(main_frame, text=f"Updated: {VERSION_DATE}", 
                         font=('Arial', 8))
version_label.grid(row=1, column=0, pady=(0, 15))

# Instructions
instructions = """• Select MULTIPLE files to compare different scenarios (hold Ctrl)
• Enable Min/Max Optimization to find the best sorting strategy for each file
• Use Quick Analysis Mode for faster results without Excel export"""

inst_label = ttk.Label(main_frame, text=instructions, justify=tk.LEFT, wraplength=700)
inst_label.grid(row=2, column=0, pady=(0, 20))

# Min/Max Mode checkbox with tooltip
minmax_frame = ttk.Frame(main_frame)
minmax_frame.grid(row=3, column=0, pady=(0, 10))

minmax_var = tk.BooleanVar()
minmax_checkbox = ttk.Checkbutton(
    minmax_frame, 
    text="🔥 Enable Triple Optimization Mode",
    variable=minmax_var,
    style='Big.TCheckbutton'
)
minmax_checkbox.grid(row=0, column=0)

minmax_tooltip = ttk.Label(
    minmax_frame,
    text="Tests all sorting strategies to find best results for Orders, Hours, and Quantity",
    font=('Arial', 8, 'italic'),
    foreground='gray'
)
minmax_tooltip.grid(row=1, column=0, pady=(0, 10))

# No Export checkbox with tooltip
no_export_frame = ttk.Frame(main_frame)
no_export_frame.grid(row=4, column=0, pady=(0, 10))

no_export_var = tk.BooleanVar()
no_export_checkbox = ttk.Checkbutton(
    no_export_frame, 
    text="⚡ Quick Analysis Mode",
    variable=no_export_var,
    style='Big.TCheckbutton'
)
no_export_checkbox.grid(row=0, column=0)

no_export_tooltip = ttk.Label(
    no_export_frame,
    text="Show results instantly without creating Excel files (useful for rapid testing)",
    font=('Arial', 8, 'italic'),
    foreground='gray'
)
no_export_tooltip.grid(row=1, column=0, pady=(0, 10))

# Process button
process_btn = ttk.Button(main_frame, text="📂 SELECT FILES & PROCESS", 
                        command=load_and_process_files, 
                        style='Big.TButton')
process_btn.grid(row=5, column=0, pady=(10, 20))

# Configure button style
style = ttk.Style()
style.configure('Big.TButton', font=('Arial', 12, 'bold'))
style.configure('Big.TCheckbutton', font=('Arial', 10, 'bold'))
style.configure('Success.TFrame', background='#7ff09a')

# Status
status_var = tk.StringVar()
status_var.set("🔄 Ready - Select Excel file(s) to begin processing")
status_label = ttk.Label(main_frame, textvariable=status_var, font=('Arial', 10))
status_label.grid(row=6, column=0, pady=(0, 10), sticky=tk.W)

# Results area
results_frame = ttk.LabelFrame(main_frame, text="📊 Results", padding="10")
results_frame.grid(row=7, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(10, 0))

results_text = tk.Text(results_frame, height=18, width=90, font=('Consolas', 9))
scrollbar = ttk.Scrollbar(results_frame, orient=tk.VERTICAL, command=results_text.yview)
results_text.configure(yscrollcommand=scrollbar.set)

results_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))

copy_btn = ttk.Button(main_frame, text="📋 Copy Summary", command=copy_summary_to_clipboard)
copy_btn.grid(row=8, column=0, pady=(10, 10))

# Configure grid weights
root.columnconfigure(0, weight=1)
root.rowconfigure(0, weight=1)
main_frame.columnconfigure(0, weight=1)
main_frame.rowconfigure(7, weight=1)
results_frame.columnconfigure(0, weight=1)
results_frame.rowconfigure(0, weight=1)

# Show initial message
results_text.insert(1.0, "Select your Excel file(s) to begin material release planning...")

if __name__ == "__main__":
    root.mainloop()